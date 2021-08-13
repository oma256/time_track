from io import BytesIO

from django.contrib.auth import authenticate, login
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.db import IntegrityError

from rest_framework.exceptions import NotFound

import xlsxwriter
from openpyxl import load_workbook

from apps.organizations.models import UserOrganization
from apps.organizations.services import (
    OrganizationSettingQueryService,
    OrganizationQueryService,
)
from apps.payment.services import OrganizationTariffPackageQueryService
from apps.services.base import BaseQueryService
from apps.users.models import User
from utils.user_phone_number_regex import clean_phone


class UserQueryService(BaseQueryService):
    model = User

    @classmethod
    def user_login(cls, request, form):
        _phone = form.cleaned_data.get('phone_number')
        _pass = form.cleaned_data.get('password')
        user = authenticate(
            request, phone_number=clean_phone(_phone), password=_pass
        )
        user_org = user.user_organizations.first()

        if user_org is None:
            return None, None

        if user is not None and user_org.organization.admin == user:
            login(request, user)
            return user, user_org
        return None, None

    @classmethod
    def user_registration(cls, request, form):
        form_data = form.cleaned_data
        org_name = form_data.pop('org_name')
        form_data.pop('password2')

        user = cls.model.objects.create_user(**form_data)
        login(request, user)

        return {'user': user, 'org_name': org_name}

    @classmethod
    def create_user(cls, data):
        data = cls.clean_phone(data)

        user, _ = cls.model.objects.get_or_create(
            phone_number=data.get('phone_number'),
            defaults={
                'last_name': data.get('last_name'),
                'first_name': data.get('first_name'),
                'middle_name': data.get('middle_name', None),
            }
        )

        return user

    @classmethod
    def update_user(cls, user, data):
        user.last_name = data.get('last_name')
        user.first_name = data.get('first_name')
        user.middle_name = data.get('middle_name')
        user.phone_number = clean_phone(data.get('phone_number'))

        try:
            user.save(update_fields=[
                'last_name', 'first_name', 'middle_name', 'phone_number',
            ])
        except IntegrityError:
            return {'detail': 'error', 'status_code': 400}

        return {'detail': 'success', 'status_code': 200}

    @classmethod
    def get_users_by_department(cls, department):
        return cls.model.objects.filter(
            user_organizations__organization=department
        )

    @classmethod
    def get_user_by_id(cls, user_id):
        try:
            return cls.model.objects.get(id=user_id)
        except cls.model.DoesNotExist:
            return None

    @classmethod
    def get_user_by_phone_number(cls, phone_number):
        try:
            return User.objects.get(phone_number=phone_number)
        except User.DoesNotExist:
            raise NotFound(detail='User not found')

    @classmethod
    def change_user_password(cls, user, form):
        user.set_password(form.cleaned_data.get('password'))
        user.save(update_fields=['password'])

    @classmethod
    def get_login_success_url(cls, user_org=None, user=None):
        if user_org:
            org = user_org.organization
            org_setting = (
                OrganizationSettingQueryService.get_setting_by_org_id(org.id)
            )

            cls._generate_url(org=org, org_setting=org_setting)

        user_org = user.user_organizations.first()
        org = user_org.organization
        org_setting = (
            OrganizationSettingQueryService.get_setting_by_org_id(org.id)
        )

        return cls._generate_url(org=org, org_setting=org_setting)

    @classmethod
    def _generate_url(cls, org, org_setting):
        org_tariff_pkg = (
            OrganizationTariffPackageQueryService.get_tariff_package_by_org(
                org=org
            )
        )
        org_tariff_pkg_is_not_payed = (
            OrganizationTariffPackageQueryService.check_org_tariff_pkg_payed(
                org_tariff_pkg=org_tariff_pkg
            )
        )
        if org_setting.location:
            if org_tariff_pkg_is_not_payed:
                return reverse(
                    'organizations:tariffs', kwargs={'org_pk': org.id},
                )
            return reverse('organizations:main', kwargs={'org_pk': org.id})
        return reverse('organizations:settings', kwargs={'org_pk': org.id})

    @classmethod
    def check_user_exist(cls, user, phone_number):
        return cls.model.objects.filter(phone_number=phone_number)

    @classmethod
    def generate_import_excel_template(cls, org, filials):
        output = BytesIO()
        org_departments = (
            OrganizationQueryService.get_organization_departments(org=org)
        )

        with xlsxwriter.Workbook(output) as workbook:
            worksheet = workbook.add_worksheet('employees import template')
            bold = workbook.add_format({'bold': True, 'border': 1})
            border = workbook.add_format({'border': 1})
            employee_info_head = (
                'Фамилия *', 'Имя *', 'Отчество', 'Номер телефона *',
                'Должность *', 'Филиал (ID)', 'Отдел (ID)',
            )
            employee_example_info = (
                'Иванов', 'Иван', 'Иванович', '996776999999', 'Программист',
                'ID', 'ID',
            )
            row_index = 0

            worksheet.write(2, 0, 'Филиалы', bold)
            worksheet.write(3, 1, 'Название', bold)
            worksheet.write(3, 2, 'Головной офис', border)
            worksheet.write(4, 1, 'ID', bold)
            worksheet.write(4, 2, org.id, border)
            worksheet.write(3, 4, 'Отделы', bold)
            if org_departments:
                worksheet.write(4, 5, 'ID', bold)
                worksheet.write(4, 6, 'Название', bold)

                for i, item in enumerate(org_departments, start=5):
                    worksheet.write(i, 5, item.id, border)
                    worksheet.write(i, 6, item.name, border)
                    row_index = i
            else:
                worksheet.write(4, 5, 'Нет отделов', border)
                row_index = 4

            if not filials:
                worksheet.write(9, 1, 'Нет филиалов', border)
            else:
                index = row_index + 3
                for filial in filials:
                    worksheet.write(index, 1, 'Название', bold)
                    worksheet.write(index, 2, filial.name, border)
                    worksheet.write(index+1, 1, 'ID', bold)
                    worksheet.write(index+1, 2, filial.id, border)
                    worksheet.write(index, 4, 'Отделы', bold)

                    departments = (
                        OrganizationQueryService.get_departments_by_filial(
                            filial=filial,
                        )
                    )
                    if not departments:
                        worksheet.write(index+1, 5, 'Нет отделов', border)
                        index += 4
                    else:
                        worksheet.write(index+1, 5, 'ID', bold)
                        worksheet.write(index+1, 6, 'Название', bold)
                        for j, depart in enumerate(departments, start=index+2):
                            worksheet.write(j, 5, depart.id, border)
                            worksheet.write(j, 6, depart.name, border)
                            index = j + 3

            for col, item in enumerate(employee_info_head, start=0):
                worksheet.write(28, col, item, bold)

            for col, item in enumerate(employee_example_info, start=0):
                worksheet.write(29, col, item, border)

        output.seek(0)

        return HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-office'
                         'document.spreadsheetml.sheet',
        )

    @classmethod
    def import_employee_data_to_db(cls, file, org):
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        def _update_user_org_params(org, **params):
            if org.is_department:
                params.update({
                    'organization': org,
                    'start_time': org.parent_org.org_settings.start_time,
                    'end_time': org.parent_org.org_settings.end_time,
                    'non_fined_minute': (
                        org.parent_org.org_settings.non_fined_minute
                    ),
                })
            else:
                params.update({
                    'organization': org,
                    'start_time': org.org_settings.start_time,
                    'end_time': org.org_settings.end_time,
                    'non_fined_minute': org.org_settings.non_fined_minute,
                })

            return params

        def _create_user(**params):
            if (params.get('last_name')
                    and params.get('first_name')
                    and params.get('phone_number')):
                try:
                    return User.objects.create(**params)
                except IntegrityError:
                    return None
            else:
                return JsonResponse({'error': 'required fields'}, status=400)

        for row in sheet.iter_rows(min_row=30, max_col=600):
            filial_id = row[5].value
            department_id = row[6].value
            user_params = {
                'last_name': row[0].value,
                'first_name': row[1].value,
                'middle_name': row[2].value,
                'phone_number': f'+{row[3].value}',
            }
            user_org_params = {
                'position': row[4].value,
            }

            user = _create_user(**user_params)

            if not user:
                continue

            user_org_params.update({'user': user})

            if department_id and user_org_params.get('position'):
                department = OrganizationQueryService.get_organization_by_id(
                    org_id=department_id,
                )
                if department is None:
                    return JsonResponse({
                        'error': 'un correct data'
                    }, status=400)

                updated_params = _update_user_org_params(
                    org=department, **user_org_params,
                )
            elif filial_id and user_org_params.get('position'):
                filial = OrganizationQueryService.get_organization_by_id(
                    org_id=filial_id,
                )

                if filial is None:
                    return JsonResponse({
                        'error': 'un correct data'
                    }, status=400)

                updated_params = _update_user_org_params(
                    org=filial, **user_org_params,
                )
            elif user_org_params.get('position'):
                updated_params = _update_user_org_params(
                    org=org, **user_org_params,
                )
            else:
                return JsonResponse({'error': 'required fields'}, status=400)

            UserOrganization.objects.create(**updated_params)

        return JsonResponse({'detail': 'created'}, status=201)
